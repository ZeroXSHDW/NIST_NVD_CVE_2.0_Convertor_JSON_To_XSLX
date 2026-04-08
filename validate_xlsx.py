"""
NIST CVE Data Integrity Validator
----------------------------------
Verifies the generated XLSX workbook against the source JSON files to
ensure 100% data accuracy and consistency.
"""

import json
import re
from pathlib import Path
from openpyxl import load_workbook

# --- Configuration ---
DATA_DIR = Path(__file__).parent / "nvd_data"
XLSX_PATH = Path(__file__).parent / 'NIST_CVE_Compiled.xlsx'

# Columns order used in the workbook
COLUMNS = ['cve_id', 'description', 'published', 'lastModified', 'baseScore', 'severity', 'vectorString']

def clean_text(text):
    if not isinstance(text, str):
        return str(text or '')
    # Normalize variants of line endings: \r\n, \n\r, \r to \n
    return text.replace('\r\n', '\n').replace('\n\r', '\n').replace('\r', '\n')

def extract_entry(cve_entry):
    cve_id = str(cve_entry.get('id', ''))
    description = ''
    for desc in cve_entry.get('descriptions', []):
        if desc.get('lang') == 'en':
            description = clean_text(desc.get('value', ''))
            break
    published = str(cve_entry.get('published', ''))
    last_modified = str(cve_entry.get('lastModified', ''))

    base_score = ''
    severity = ''
    vector = ''
    metrics = cve_entry.get('metrics', {})

    # Ordered list of standards to check
    metric_keys = ['cvssMetricV40', 'cvssMetricV31', 'cvssMetricV30', 'cvssMetricV2']

    found_metric = None
    # First priority: Primary metric
    for key in metric_keys:
        for m in metrics.get(key, []):
            if m.get('type') == 'Primary':
                found_metric = (key, m)
                break
        if found_metric:
            break

    # Second priority: Any metric (Secondary/etc.)
    if not found_metric:
        for key in metric_keys:
            for m in metrics.get(key, []):
                found_metric = (key, m)
                break
            if found_metric:
                break

    if found_metric:
        key, m = found_metric
        cvss_data = m.get('cvssData', {})
        ds = cvss_data.get('baseScore', '')
        base_score = ds if ds is not None else ''
        severity = cvss_data.get('baseSeverity', '') or m.get('baseSeverity', '') or m.get('severity', '') or ''
        vector = cvss_data.get('vectorString', '') or ''

    return {
        'cve_id': cve_id,
        'description': description,
        'published': published,
        'lastModified': last_modified,
        'baseScore': base_score,
        'severity': str(severity),
        'vectorString': str(vector),
    }

def normalize(val):
    if val is None or val == "":
        return ""
    # Normalize float scores for consistent comparison
    try:
        f = float(val)
        return str(f)
    except (TypeError, ValueError):
        # Normalize line endings
        return clean_text(val)

def main():
    print(f"Loading workbook {XLSX_PATH} for definitive perfection check...")
    wb = load_workbook(XLSX_PATH, read_only=True)
    mismatches = []
    json_files = sorted(DATA_DIR.glob('nvdcve-2.0-*.json'))
    for json_path in json_files:
        year_match = re.search(r'nvdcve-2\.0-(\d{4})\.json', json_path.name)
        if not year_match:
            continue
        year = year_match.group(1)
        ws = wb[year]
        print(f"Checking {year}...")
        with open(json_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        vulns = json_data.get('vulnerabilities', [])
        # Build lookup
        json_lookup = {v['cve']['id']: extract_entry(v['cve']) for v in vulns}
        # Iterate rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(COLUMNS, row))
            cve_id = row_dict['cve_id']
            if cve_id not in json_lookup:
                mismatches.append((year, cve_id, 'Missing in JSON'))
                continue
            json_entry = json_lookup[cve_id]
            for key in COLUMNS:
                wb_val = normalize(row_dict[key])
                js_val = normalize(json_entry[key])
                if wb_val != js_val:
                    mismatches.append((year, cve_id, key, wb_val, js_val))
    if mismatches:
        print('Found mismatches:')
        for m in mismatches[:10]:
            print(m)
        print(f"Total mismatches: {len(mismatches)}")
    else:
        print('All rows match source JSON perfectly.')

if __name__ == '__main__':
    main()
