"""
NVD CVE JSON to XLSX Converter
------------------------------
Processes a collection of NVD CVE 2.0 JSON files and compiles them into
a single, highly-structured Excel workbook with annual sheets and a master index.
"""

import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# --- Configuration ---
DATA_DIR = Path(__file__).parent / "nvd_data"
OUTPUT_FILE = Path(__file__).parent / 'NIST_CVE_Compiled.xlsx'

# Define columns we want in the annual sheets
COLUMNS = [
    ('cve_id', 'CVE ID'),
    ('description', 'Description (EN)'),
    ('published', 'Published Date'),
    ('lastModified', 'Last Modified Date'),
    ('baseScore', 'Base Score'),
    ('severity', 'Severity'),
    ('vectorString', 'CVSS Vector'),
]

# Columns for the Master INDEX sheet
INDEX_COLUMNS = [
    ('cve_id', 'CVE ID'),
    ('year', 'Year/Sheet'),
    ('severity', 'Severity'),
    ('baseScore', 'Base Score'),
    ('summary', 'Summary (Short Description)'),
]

def clean_text(text):
    if not isinstance(text, str):
        return str(text or '')
    return text.replace('\r\n', '\n').replace('\n\r', '\n').replace('\r', '\n')

def extract_entry(cve_entry):
    cve_id = str(cve_entry.get('id', ''))
    description = ''
    for desc in cve_entry.get('descriptions', []):
        if desc.get('lang') == 'en':
            description = clean_text(desc.get('value', ''))
            break
    published = str(cve_entry.get('published', ''))
    last_modified = str(cve_entry.get('lastModified', ''))

    base_score = ''
    severity = ''
    vector = ''
    metrics = cve_entry.get('metrics', {})

    metric_keys = ['cvssMetricV40', 'cvssMetricV31', 'cvssMetricV30', 'cvssMetricV2']
    found_metric = None
    for key in metric_keys:
        for m in metrics.get(key, []):
            if m.get('type') == 'Primary':
                found_metric = (key, m)
                break
        if found_metric: break

    if not found_metric:
        for key in metric_keys:
            for m in metrics.get(key, []):
                found_metric = (key, m)
                break
            if found_metric: break

    if found_metric:
        key, m = found_metric
        cvss_data = m.get('cvssData', {})
        ds = cvss_data.get('baseScore', '')
        base_score = ds if ds is not None else ''
        severity = cvss_data.get('baseSeverity', '') or m.get('baseSeverity', '') or m.get('severity', '') or ''
        vector = cvss_data.get('vectorString', '') or ''

    return {
        'cve_id': cve_id,
        'description': description,
        'published': published,
        'lastModified': last_modified,
        'baseScore': base_score,
        'severity': str(severity),
        'vectorString': str(vector),
    }

def main():
    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # We will accumulate index entries here
    master_index = []

    json_files = sorted(DATA_DIR.glob('nvdcve-2.0-*.json'))
    for json_path in json_files:
        match = re.search(r'nvdcve-2\.0-(\d{4})\.json', json_path.name)
        if not match: continue
        year = match.group(1)
        print(f"Processing {year}...")
        ws = wb.create_sheet(title=year)
        # Headers
        for col_idx, (_, header) in enumerate(COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        vulns = data.get('vulnerabilities', [])
        row_num = 2
        for vuln in vulns:
            cve_entry = vuln.get('cve', {})
            extracted = extract_entry(cve_entry)
            # Write to annual sheet
            for col_idx, (key, _) in enumerate(COLUMNS, start=1):
                ws.cell(row=row_num, column=col_idx, value=extracted.get(key, ''))
            
            # Add to master index
            summary = extracted['description'][:200] + ('...' if len(extracted['description']) > 200 else '')
            master_index.append({
                'cve_id': extracted['cve_id'],
                'year': year,
                'severity': extracted['severity'],
                'baseScore': extracted['baseScore'],
                'summary': summary
            })
            row_num += 1
        # Auto-adjust annual sheet widths
        for col_idx, (key, header) in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header), 15)

    # Create Master INDEX sheet at the front
    print("Creating Master INDEX sheet...")
    idx_ws = wb.create_sheet(title='INDEX', index=0)
    for col_idx, (_, header) in enumerate(INDEX_COLUMNS, start=1):
        idx_ws.cell(row=1, column=col_idx, value=header)
    
    # Sort index by CVE ID (Reverse chronological usually best for lookups)
    master_index.sort(key=lambda x: x['cve_id'], reverse=True)
    
    for row_num, entry in enumerate(master_index, start=2):
        for col_idx, (key, _) in enumerate(INDEX_COLUMNS, start=1):
            idx_ws.cell(row=row_num, column=col_idx, value=entry.get(key, ''))
    
    # Auto-adjust INDEX widths
    for col_idx, (key, header) in enumerate(INDEX_COLUMNS, start=1):
        idx_ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header), 15)

    wb.save(OUTPUT_FILE)
    print(f'Workbook saved with INDEX to {OUTPUT_FILE}')

if __name__ == '__main__':
    main()
